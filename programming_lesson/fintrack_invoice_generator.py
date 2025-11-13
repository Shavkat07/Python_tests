from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def ensure_invoices_dir(base: Path | str = "./invoices") -> Path:
    path = Path(base)
    path.mkdir(parents=True, exist_ok=True)
    return path


def timestamp_str(now: Optional[datetime] = None) -> str:
    now = now or datetime.now()
    return now.strftime("%Y%m%d_%H%M%S")


class InvoiceGenerator(ABC):
    def __init__(self, client_name: str, items: List[Dict[str, float]]):
        self.client_name = client_name
        self.items = items
        self.created_at = datetime.now()

    def calculate_total(self) -> float:
        total = 0.0
        for item in self.items:
            try:
                total += float(item.get("price", 0.0))
            except Exception:
                raise ValueError(f"Invalid price: {item}")
        return round(total, 2)

    @abstractmethod
    def generate_invoice(self, output_dir: Path | str = "./invoices") -> Path:
        pass



class PDFInvoiceGenerator(InvoiceGenerator):
    def generate_invoice(self, output_dir: Path | str = "./invoices") -> Path:
        outdir = ensure_invoices_dir(output_dir)
        filename = f"invoice_pdf_{self.client_name}_{timestamp_str(self.created_at)}.pdf"
        path = outdir / filename

        c = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4

        margin = 20 * mm
        y = height - margin

        c.setFont("Helvetica-Bold", 16)
        c.drawString(margin, y, f"Invoice — {self.client_name}")
        y -= 12 * mm

        c.setFont("Helvetica", 9)
        c.drawString(margin, y, f"Created: {self.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
        y -= 8 * mm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin, y, "Item")
        c.drawString(width - margin - 60 * mm, y, "Price")
        y -= 6 * mm

        c.setFont("Helvetica", 10)
        for item in self.items:
            c.drawString(margin, y, item.get("name", ""))
            c.drawRightString(width - margin, y, f"{float(item.get('price', 0)):.2f}")
            y -= 6 * mm
            if y < margin + 30 * mm:
                c.showPage()
                y = height - margin

        total = self.calculate_total()
        y -= 6 * mm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, "Total:")
        c.drawRightString(width - margin, y, f"{total:.2f}")

        c.showPage()
        c.save()
        return path


class ExcelInvoiceGenerator(InvoiceGenerator):
    def generate_invoice(self, output_dir: Path | str = "./invoices") -> Path:
        outdir = ensure_invoices_dir(output_dir)
        filename = f"invoice_xlsx_{self.client_name}_{timestamp_str(self.created_at)}.xlsx"
        path = outdir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        ws["A1"], ws["B1"] = "Item", "Price"
        row = 2
        for item in self.items:
            ws[f"A{row}"] = item.get("name", "")
            ws[f"B{row}"] = float(item.get("price", 0))
            row += 1

        ws[f"A{row}"] = "Total"
        ws[f"B{row}"] = self.calculate_total()
        row += 1
        ws[f"A{row}"] = "Created"
        ws[f"B{row}"] = self.created_at.strftime('%Y-%m-%d %H:%M:%S')

        for col in range(1, 3):
            col_letter = get_column_letter(col)
            max_len = max(len(str(cell.value)) for cell in ws[col_letter] if cell.value)
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(path)
        return path


class HTMLInvoiceGenerator(InvoiceGenerator):
    def generate_invoice(self, output_dir: Path | str = "./invoices") -> Path:
        outdir = ensure_invoices_dir(output_dir)
        filename = f"invoice_html_{self.client_name}_{timestamp_str(self.created_at)}.html"
        path = outdir / filename

        total = self.calculate_total()
        created = self.created_at.strftime('%Y-%m-%d %H:%M:%S')

        html = [
            "<!doctype html>",
            "<html lang='en'>",
            "<head>",
            f"<title>Invoice — {self.client_name}</title>",
            "<meta charset='utf-8'>",
            "<style>body{font-family:Arial,sans-serif;max-width:800px;margin:2rem auto;}table{width:100%;border-collapse:collapse;}th,td{border:1px solid #ccc;padding:6px;text-align:left;}</style>",
            "</head>",
            "<body>",
            f"<h1>Invoice — {self.client_name}</h1>",
            f"<p>Created: {created}</p>",
            "<table><tr><th>Item</th><th>Price</th></tr>",
        ]

        for item in self.items:
            html.append(f"<tr><td>{item.get('name','')}</td><td>{float(item.get('price',0)):.2f}</td></tr>")

        html += [
            "</table>",
            f"<p><b>Total:</b> {total:.2f}</p>",
            "</body></html>",
        ]

        path.write_text("\n".join(html), encoding="utf-8")
        return path


class InvoiceManager:
    def __init__(self, generator: InvoiceGenerator):
        self.generator = generator

    def export_invoice(self, output_dir: Path | str = "./invoices") -> Path:
        path = self.generator.generate_invoice(output_dir)
        print(f"Invoice created: {path}")
        return path


def generator_factory(format_name: str, client_name: str, items: List[Dict[str, float]]) -> InvoiceGenerator:
    fmt = format_name.lower()
    if fmt == "pdf":
        return PDFInvoiceGenerator(client_name, items)
    if fmt in ("excel", "xlsx", "xls"):
        return ExcelInvoiceGenerator(client_name, items)
    if fmt in ("html", "htm"):
        return HTMLInvoiceGenerator(client_name, items)
    raise ValueError(f"Unsupported format: {format_name}")


if __name__ == "__main__":
    demo_items = [
        {"name": "Backend development (10h)", "price": 300.0},
        {"name": "Design review", "price": 75.5},
        {"name": "Hosting reimbursement", "price": 20.0},
    ]

    outdir = ensure_invoices_dir()

    pdf_gen = generator_factory("pdf", "PDP University", demo_items)
    xlsx_gen = generator_factory("xlsx", "PDP University", demo_items)
    html_gen = generator_factory("html", "PDP University", demo_items)

    for gen in [pdf_gen, xlsx_gen, html_gen]:
        manager = InvoiceManager(gen)
        manager.export_invoice(outdir)
